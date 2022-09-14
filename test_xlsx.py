from openpyxl import load_workbook
workbook = load_workbook('resources/file_example_XLSX_50.xlsx')
sheet = workbook.active
print(sheet.cell(row=3, column=2).value)
for x in range(1, 10):
    for y in range(1, 10):
        print(sheet.cell(row=x, column=y).value)

for row in sheet.values:
    for value in row:
        print(value)
